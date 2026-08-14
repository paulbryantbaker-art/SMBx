#!/usr/bin/env python3
"""Build model.xlsx from a runner result.json.

Real Excel formulas — the workbook recalculates when inputs change.
IB conventions: blue inputs, black formulas, yellow key assumptions,
$#,##0 currency, negatives in parentheses, multiples as 0.0x.
Engine outputs from result.json are cross-checked against the formula cells
after recalc by verify_xlsx.py (parity: Excel must agree with the engine).
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = Font(name="Arial", color="0000FF")
BLACK = Font(name="Arial")
BOLD = Font(name="Arial", bold=True)
H1 = Font(name="Arial", bold=True, size=14)
YELLOW = PatternFill("solid", fgColor="FFFF00")
MONEY = "$#,##0;($#,##0);-"
PCT = "0.0%"
MULT = "0.00x"
RATIO = "0.00"
THIN = Border(bottom=Side(style="thin", color="EEEEF0"))

def sheet_setup(ws, title, widths=(34, 16, 16, 16, 16, 16, 16, 16)):
    ws["A1"] = title
    ws["A1"].font = H1
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w

def put(ws, cell, value, font=BLACK, fmt=None, fill=None):
    ws[cell] = value
    ws[cell].font = font
    if fmt: ws[cell].number_format = fmt
    if fill: ws[cell].fill = fill

def legend(ws, row):
    put(ws, f"A{row}", "Legend: blue = input (edit these) · black = formula · yellow = key assumption", Font(name="Arial", italic=True, size=9))

def build_earnings_sheet(wb, deal):
    """Assumptions + Earnings: SDE and Adjusted EBITDA per V17 §5.1/§5.2, as formulas."""
    ws = wb.create_sheet("Earnings")
    sheet_setup(ws, f"Earnings Normalization — {deal['deal'].get('business', deal['deal']['name'])}")
    years = deal["financials"]["years"]
    rows = [("Revenue", "revenue"), ("Net income", "net_income"), ("Owner salary", "owner_salary"),
            ("Depreciation", "depreciation"), ("Amortization", "amortization"), ("Interest", "interest"),
            ("Taxes", "taxes"), ("One-time expenses", "one_time_expenses"),
            ("Non-recurring income", "non_recurring_income")]
    r0 = 3
    put(ws, f"A{r0}", "Line item", BOLD)
    for j, y in enumerate(years):
        put(ws, f"{chr(66+j)}{r0}", str(y["label"]), BOLD)
    for i, (label, key) in enumerate(rows):
        r = r0 + 1 + i
        put(ws, f"A{r}", label)
        for j, y in enumerate(years):
            put(ws, f"{chr(66+j)}{r}", y.get(key, 0) or 0, BLUE, MONEY)
    # verified add-backs, one row (sum of verified items per year)
    ab_row = r0 + 1 + len(rows)
    put(ws, f"A{ab_row}", "Verified add-backs (schedule below)")
    addbacks = [a for a in deal["financials"].get("addbacks", []) if a.get("verified")]
    for j, y in enumerate(years):
        amt = sum(a["amount"] for a in addbacks if not a.get("year") or a.get("year") == y["label"])
        put(ws, f"{chr(66+j)}{ab_row}", amt, BLUE, MONEY)
    # SDE = NI + owner salary + D + A + interest + one-time + verified addbacks  (V17 §5.1)
    sde_row, ebd_row = ab_row + 2, ab_row + 3
    put(ws, f"A{sde_row}", "SDE (V17 §5.1)", BOLD)
    put(ws, f"A{ebd_row}", "Adjusted EBITDA (V17 §5.2)", BOLD)
    for j in range(len(years)):
        c = chr(66 + j)
        ni, osal, dep, amo, intr, tax, ote, nri = (f"{c}{r0+1+i}" for i in (1, 2, 3, 4, 5, 6, 7, 8))
        ab = f"{c}{ab_row}"
        put(ws, f"{c}{sde_row}", f"={ni}+{osal}+{dep}+{amo}+{intr}+{ote}+{ab}", BOLD, MONEY)
        put(ws, f"{c}{ebd_row}", f"={ni}+{dep}+{amo}+{intr}+{tax}+{ab}-{nri}", BOLD, MONEY)
    # add-back schedule with evidence
    r = ebd_row + 2
    put(ws, f"A{r}", "Add-back schedule (verified only enter the math)", BOLD)
    r += 1
    for h, ccol in (("Item", "A"), ("Year", "B"), ("Amount", "C"), ("Verified", "D"), ("Evidence", "E")):
        put(ws, f"{ccol}{r}", h, BOLD)
    for a in deal["financials"].get("addbacks", []):
        r += 1
        put(ws, f"A{r}", a["label"]); put(ws, f"B{r}", a.get("year", "all"))
        put(ws, f"C{r}", a["amount"], BLUE, MONEY)
        put(ws, f"D{r}", "YES" if a.get("verified") else "NO — excluded")
        put(ws, f"E{r}", a.get("evidence", ""))
    legend(ws, r + 2)
    last_col = chr(65 + len(years))
    return {"sde_cell": f"Earnings!{last_col}{sde_row}", "ebitda_cell": f"Earnings!{last_col}{ebd_row}"}

def build_valuation(wb, res):
    deal = res["inputs"]
    refs = build_earnings_sheet(wb, deal)
    ws = wb.create_sheet("Valuation")
    sheet_setup(ws, "Valuation Range")
    v = deal.get("valuation", {})
    basis = (v.get("multiple_basis") or res["current"]["basis"])
    basis_ref = refs["sde_cell"] if basis == "SDE" else refs["ebitda_cell"]
    put(ws, "A3", f"Earnings basis ({basis}, latest year)"); put(ws, "B3", f"={basis_ref}", BOLD, MONEY)
    put(ws, "A4", "Multiple — low");  put(ws, "B4", v.get("multiple_low", 0), BLUE, MULT, YELLOW)
    put(ws, "A5", "Multiple — high"); put(ws, "B5", v.get("multiple_high", 0), BLUE, MULT, YELLOW)
    put(ws, "A6", "Multiple source"); put(ws, "B6", v.get("multiple_source", "UNSOURCED — do not publish"))
    put(ws, "A8", "Value — low", BOLD);  put(ws, "B8", "=B3*B4", BOLD, MONEY)
    put(ws, "A9", "Value — mid", BOLD);  put(ws, "B9", "=B3*(B4+B5)/2", BOLD, MONEY)
    put(ws, "A10", "Value — high", BOLD); put(ws, "B10", "=B3*B5", BOLD, MONEY)
    legend(ws, 12)

def loan_rows(deal):
    s = deal.get("structure", {})
    loans = []
    if s.get("sba_loan"):
        loans.append(("SBA 7(a)", s["sba_loan"], s.get("sba_rate", 0), s.get("sba_term_years", 10)))
    if s.get("seller_note"):
        loans.append(("Seller note", s["seller_note"], s.get("seller_note_rate", 0), s.get("seller_note_term_years", 5)))
    return loans

def build_sba(wb, res):
    deal = res["inputs"]
    refs = build_earnings_sheet(wb, deal)
    ws = wb.create_sheet("SBA")
    sheet_setup(ws, "Capital Stack & DSCR")
    s = deal.get("structure", {})
    put(ws, "A3", "Purchase price"); put(ws, "B3", s.get("purchase_price", 0), BLUE, MONEY, YELLOW)
    loans = loan_rows(deal)
    r = 5
    put(ws, f"A{r}", "Loan", BOLD); put(ws, f"B{r}", "Principal", BOLD); put(ws, f"C{r}", "Rate", BOLD)
    put(ws, f"D{r}", "Term (yrs)", BOLD); put(ws, f"E{r}", "Monthly pmt", BOLD); put(ws, f"F{r}", "Annual service", BOLD)
    first, last = r + 1, r + len(loans)
    for i, (name, prin, rate, term) in enumerate(loans):
        rr = r + 1 + i
        put(ws, f"A{rr}", name)
        put(ws, f"B{rr}", prin, BLUE, MONEY); put(ws, f"C{rr}", rate, BLUE, PCT); put(ws, f"D{rr}", term, BLUE)
        # PMT: V17 §4.5.4 mortgage amortization == Excel PMT on monthly r, n
        put(ws, f"E{rr}", f"=IF(C{rr}=0,B{rr}/(D{rr}*12),-PMT(C{rr}/12,D{rr}*12,B{rr}))", BLACK, MONEY)
        put(ws, f"F{rr}", f"=E{rr}*12", BLACK, MONEY)
    tr = last + 1
    put(ws, f"A{tr}", "Total debt / annual service", BOLD)
    put(ws, f"B{tr}", f"=SUM(B{first}:B{last})", BOLD, MONEY)
    put(ws, f"F{tr}", f"=SUM(F{first}:F{last})", BOLD, MONEY)
    r = tr + 2
    put(ws, f"A{r}", "Equity injection"); put(ws, f"B{r}", f"=B3-B{tr}", BLACK, MONEY)
    put(ws, f"A{r+1}", "LTV"); put(ws, f"B{r+1}", f"=IF(B3=0,0,B{tr}/B3)", BLACK, PCT)
    put(ws, f"A{r+3}", "EBITDA (latest)"); put(ws, f"B{r+3}", f"={refs['ebitda_cell']}", BLACK, MONEY)
    put(ws, f"A{r+4}", "SDE (latest)"); put(ws, f"B{r+4}", f"={refs['sde_cell']}", BLACK, MONEY)
    put(ws, f"A{r+5}", "DSCR on EBITDA (V17 §5.3)", BOLD); put(ws, f"B{r+5}", f"=IF(F{tr}=0,0,B{r+3}/F{tr})", BOLD, RATIO)
    put(ws, f"A{r+6}", "DSCR on SDE", BOLD); put(ws, f"B{r+6}", f"=IF(F{tr}=0,0,B{r+4}/F{tr})", BOLD, RATIO)
    put(ws, f"A{r+7}", "SBA test (≥ 1.25)"); put(ws, f"B{r+7}", f'=IF(B{r+5}>=1.25,"PASS","FAIL")', BOLD)
    put(ws, f"A{r+8}", "Conventional test (≥ 1.50)"); put(ws, f"B{r+8}", f'=IF(B{r+5}>=1.5,"PASS","FAIL")', BOLD)
    legend(ws, r + 10)
    # Amortization sheet (annual, from engine schedule — values, formula-checked by PMT above)
    wa = wb.create_sheet("Amortization")
    sheet_setup(wa, "Amortization (annual rollup — from engine schedule)")
    rr = 3
    for li, loan in enumerate(res.get("loans", [])):
        put(wa, f"A{rr}", loan["name"], BOLD); rr += 1
        for h, cc in (("Year", "A"), ("Interest", "B"), ("Principal", "C"), ("Balance", "D")):
            put(wa, f"{cc}{rr}", h, BOLD)
        for row in loan["schedule"]:
            rr += 1
            put(wa, f"A{rr}", row["year"])
            put(wa, f"B{rr}", round(row["interest"], 2), BLACK, MONEY)
            put(wa, f"C{rr}", round(row["principal"], 2), BLACK, MONEY)
            put(wa, f"D{rr}", round(row["balance"], 2), BLACK, MONEY)
        rr += 2

def build_lbo(wb, res):
    deal = res["inputs"]
    refs = build_earnings_sheet(wb, deal)
    ws = wb.create_sheet("LBO")
    sheet_setup(ws, "LBO / Returns")
    a = res["assumptions"]
    y0 = deal["financials"]["years"][-1]
    put(ws, "A3", "Revenue (latest actual)"); put(ws, "B3", y0["revenue"], BLUE, MONEY)
    put(ws, "A4", "Revenue growth"); put(ws, "B4", a["revenue_growth_pct"], BLUE, PCT, YELLOW)
    put(ws, "A5", "EBITDA margin"); put(ws, "B5", a["margin_pct"], BLUE, PCT, YELLOW)
    put(ws, "A6", "Capex % revenue"); put(ws, "B6", a["capex_pct_revenue"], BLUE, PCT)
    put(ws, "A7", "Hold (years)"); put(ws, "B7", a["hold_years"], BLUE)
    put(ws, "A8", "Entry multiple"); put(ws, "B8", a["entry_multiple"], BLACK, MULT)
    put(ws, "A9", "Exit multiple"); put(ws, "B9", a["exit_multiple"], BLUE, MULT, YELLOW)
    put(ws, "A10", "Equity invested"); put(ws, "B10", a["equity_invested"], BLACK, MONEY)
    hold = int(a["hold_years"])
    r0 = 12
    headers = ["", *[f"Year {i+1}" for i in range(hold)]]
    for j, h in enumerate(headers):
        put(ws, f"{chr(65+j)}{r0}", h, BOLD)
    put(ws, f"A{r0+1}", "Revenue")
    put(ws, f"A{r0+2}", "EBITDA")
    put(ws, f"A{r0+3}", "Capex")
    put(ws, f"A{r0+4}", "Debt service (engine sched.)")
    put(ws, f"A{r0+5}", "FCF to equity")
    put(ws, f"A{r0+6}", "DSCR")
    for y in range(1, hold + 1):
        c = chr(65 + y)
        prev = "B3" if y == 1 else f"{chr(64+y)}{r0+1}"
        put(ws, f"{c}{r0+1}", f"={prev}*(1+$B$4)", BLACK, MONEY)
        put(ws, f"{c}{r0+2}", f"={c}{r0+1}*$B$5", BLACK, MONEY)
        put(ws, f"{c}{r0+3}", f"={c}{r0+1}*$B$6", BLACK, MONEY)
        ds = res["proForma"][y-1]["debt_service"]
        put(ws, f"{c}{r0+4}", round(ds, 2), BLACK, MONEY)  # engine amortization; source: result.json
        put(ws, f"{c}{r0+5}", f"={c}{r0+2}-{c}{r0+3}-{c}{r0+4}", BLACK, MONEY)
        put(ws, f"{c}{r0+6}", f"=IF({c}{r0+4}=0,0,{c}{r0+2}/{c}{r0+4})", BLACK, RATIO)
    ex = chr(65 + hold)
    r = r0 + 8
    rem = res["proForma"][-1]["remaining_debt"]
    put(ws, f"A{r}", "Exit EV (EBITDA × exit mult)"); put(ws, f"B{r}", f"={ex}{r0+2}*$B$9", BLACK, MONEY)
    put(ws, f"A{r+1}", "Remaining debt at exit (engine)"); put(ws, f"B{r+1}", round(rem, 2), BLACK, MONEY)
    put(ws, f"A{r+2}", "Exit equity proceeds"); put(ws, f"B{r+2}", f"=B{r}-B{r+1}", BLACK, MONEY)
    # equity flow row for IRR
    fr = r + 4
    put(ws, f"A{fr}", "Equity flows", BOLD)
    put(ws, f"B{fr}", "=-B10", BLACK, MONEY)
    for y in range(1, hold + 1):
        c = chr(66 + y)
        base = f"{chr(65+y)}{r0+5}"
        put(ws, f"{c}{fr}", f"={base}" + (f"+B{r+2}" if y == hold else ""), BLACK, MONEY)
    endc = chr(66 + hold)
    put(ws, f"A{fr+2}", "IRR", BOLD); put(ws, f"B{fr+2}", f"=IRR(B{fr}:{endc}{fr})", BOLD, PCT)
    put(ws, f"A{fr+3}", "MOIC", BOLD); put(ws, f"B{fr+3}", f"=SUM(C{fr}:{endc}{fr})/B10", BOLD, MULT)
    legend(ws, fr + 5)
    # Sensitivity sheet — engine-computed IRR grid (values; two-var data tables aren't portable)
    sens = res.get("sensitivity")
    if sens:
        wsn = wb.create_sheet("Sensitivity")
        sheet_setup(wsn, "IRR — revenue growth × exit multiple (engine grid)")
        put(wsn, "A3", "growth \\ exit", BOLD)
        for j, x in enumerate(sens["exits"]):
            put(wsn, f"{chr(66+j)}3", x, BOLD, MULT)
        for i, row in enumerate(sens["rows"]):
            put(wsn, f"A{4+i}", row["growth"], BOLD, PCT)
            for j, cell in enumerate(row["cells"]):
                put(wsn, f"{chr(66+j)}{4+i}", round(cell["irr"], 4), BLACK, PCT)
        put(wsn, "A11", "Source: engine sensitivity grid (result.json). Values, not formulas — regenerate by re-running the model.", Font(name="Arial", italic=True, size=9))

def build_earnout(wb, res):
    ws = wb.create_sheet("Earnout")
    sheet_setup(ws, "Earnout — probability-weighted")
    put(ws, "A3", "Base price"); put(ws, "B3", res["base"], BLUE, MONEY)
    r = 5
    for h, cc in (("Tranche", "A"), ("Amount", "B"), ("Probability", "C"), ("Expected", "D"), ("Trigger", "E")):
        put(ws, f"{cc}{r}", h, BOLD)
    first = r + 1
    for t in res["tranches"]:
        r += 1
        put(ws, f"A{r}", t["label"]); put(ws, f"B{r}", t["amount"], BLUE, MONEY)
        put(ws, f"C{r}", t["probability"], BLUE, PCT, YELLOW)
        put(ws, f"D{r}", f"=B{r}*C{r}", BLACK, MONEY)
        put(ws, f"E{r}", t.get("trigger", ""))
    put(ws, f"A{r+2}", "Expected total price", BOLD)
    put(ws, f"B{r+2}", f"=B3+SUM(D{first}:D{r})", BOLD, MONEY)
    put(ws, f"A{r+3}", "Max total price", BOLD)
    put(ws, f"B{r+3}", f"=B3+SUM(B{first}:B{r})", BOLD, MONEY)
    legend(ws, r + 5)

def build_compare(wb, res):
    ws = wb.create_sheet("Comparison")
    sheet_setup(ws, "Deal Comparison", widths=(28,) + (18,) * 7)
    rows = [("League", "league", None), ("Earnings basis", None, None), ("Basis amount", None, MONEY),
            ("Valuation low", None, MONEY), ("Valuation high", None, MONEY),
            ("Purchase price", "purchase_price", MONEY), ("DSCR (EBITDA)", "dscr", RATIO),
            ("Risk", "risk", None), ("IRR", "irr", PCT), ("MOIC", "moic", MULT)]
    put(ws, "A3", "Metric", BOLD)
    for j, d in enumerate(res["deals"]):
        put(ws, f"{chr(66+j)}3", d["name"], BOLD)
    for i, (label, key, fmt) in enumerate(rows):
        r = 4 + i
        put(ws, f"A{r}", label)
        for j, d in enumerate(res["deals"]):
            c = f"{chr(66+j)}{r}"
            if label == "Earnings basis": put(ws, c, d["earnings"]["basis"])
            elif label == "Basis amount": put(ws, c, d["earnings"]["basisAmount"], BLACK, MONEY)
            elif label == "Valuation low": put(ws, c, (d.get("valuation_range") or {}).get("low"), BLACK, MONEY)
            elif label == "Valuation high": put(ws, c, (d.get("valuation_range") or {}).get("high"), BLACK, MONEY)
            else:
                v = d.get(key)
                put(ws, c, v if v is not None else "—", BLACK, fmt)
    put(ws, "A16", "Source: per-deal result.json runs. Values — re-run compare to refresh.", Font(name="Arial", italic=True, size=9))

def main():
    result_path, out_path = sys.argv[1], sys.argv[2]
    res = json.load(open(result_path))
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    sheet_setup(cover, "smbX deal math workbench")
    put(cover, "A3", "Model"); put(cover, "B3", res["model"], BOLD)
    put(cover, "A4", "Run date"); put(cover, "B4", res["run"]["date"])
    put(cover, "A5", "Engine"); put(cover, "B5", res["run"]["engine"])
    put(cover, "A6", "Formulas"); put(cover, "B6", "METHODOLOGY_V17 §5.0 — formulas are never invented")
    put(cover, "A8", "Blue cells are inputs; edit them and the workbook recalculates.", Font(name="Arial", italic=True, size=9))
    m = res["model"]
    if m == "valuation": build_valuation(wb, res)
    elif m == "sba": build_sba(wb, res)
    elif m == "lbo": build_lbo(wb, res)
    elif m == "earnout": build_earnout(wb, res)
    elif m == "compare": build_compare(wb, res)
    wb.save(out_path)

if __name__ == "__main__":
    main()
